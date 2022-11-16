from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
# Create your views here.
import nltk 
# from nltk.tokenize import word_tokenize 
from nltk import word_tokenize
# nltk.download('punkt')
import docx2txt
from pdfminer3.layout import LAParams, LTTextBox
from pdfminer3.pdfpage import PDFPage
from pdfminer3.pdfinterp import PDFResourceManager
from pdfminer3.pdfinterp import PDFPageInterpreter
from pdfminer3.converter import PDFPageAggregator
from pdfminer3.converter import TextConverter
import io
import os
from .models import *
# from transformers import DistilBertTokenizer, DistilBertForQuestionAnswering
import torch
from transformers import pipeline
from transformers import AutoTokenizer, AutoModelForQuestionAnswering, QuestionAnsweringPipeline


question_answerer = pipeline('question-answering',model='distilbert-base-cased-distilled-squad')

def index(request):
    return render(request,'findtexttemplate/index.html')

media_path="/var/www/coeapps/media/findtextapp/"
# media_path="findtextapp/"
def home(request):
    
    matchwords=""
    if request.method == 'POST':
        if 'searchinput' in request.POST:    
            searchinput = request.POST["searchinput"]
        if 'corpus' in request.POST: 
            corpus = request.POST["corpus"]
        if 'file' in request.FILES: 
            myfiles = request.FILES['file']
            fs = FileSystemStorage()
            filename=fs.save(media_path+myfiles.name, myfiles)
            sfiles=str(myfiles)
            # corpus=call_fileFormat("media/"+media_path+sfiles)
            corpus=call_fileFormat(media_path+sfiles)
            corpus=str(corpus)
            print("your similliar pronunciation words are--->")
    
        matchwords=matchword(corpus,searchinput)

        # matchwords.append(tmatchword)
        # return render(request,'findtexttemplate/index.html',{'matchwords':matchwords})
    return render(request,'findtexttemplate/index.html',{'matchwords':matchwords})



def QAapp(request):
    
    matchwords=""
    if request.method == 'POST':
        if 'searchinput' in request.POST:    
            searchinput = request.POST["searchinput"]
        if 'corpus' in request.POST: 
            corpus = request.POST["corpus"]
        if 'file' in request.FILES: 
            myfiles = request.FILES['file']
            fs = FileSystemStorage()
            filename=fs.save(media_path+myfiles.name, myfiles)
            sfiles=str(myfiles)
            # corpus=call_fileFormat("media/"+media_path+sfiles)
            corpus=call_fileFormat(media_path+sfiles)
            corpus=str(corpus)
            print("your similliar pronunciation words are--->")
    
        matchwords,score=getqa(corpus,searchinput)

        # matchwords.append(tmatchword)
        return render(request,'findtexttemplate/qaapp.html',{'matchwords':matchwords,'score':score})
    return render(request,'findtexttemplate/qaapp.html')



def sentapp(request):
    output=[{'label': 'POSITIVE', 'score': 0.9998656511306763}]
    iconim=""
    labeltext=""
    textcolor=""
    textbgcolor=""
    text="i love you"
    for i in output:
        label=i['label']
        accscore=i['score']
        # score+=str(accscore)
        print(label)
        labeltext+=label
        if label=="POSITIVE":
            icon="smile-icon.png"
            color="#2da007"
            bgcolor="#d7f0f0"
        elif label== "NEGATIVE":
            icon="sad-icon.png"
            color="#d10f20"
            bgcolor="#f5e1e1"
        else:
            icon="surprised-icon.png"
            color="#f5bf2d"
            bgcolor="#e3e2bb"
        # iconim.append(icon)
        
        iconim += icon
        textcolor += color
        textbgcolor += bgcolor
    context={
        'iconim':iconim,
        'labeltext':labeltext,
        'text':text,
        'textcolor':textcolor,
        'textbgcolor':textbgcolor
        
    }
    return render(request,'findtexttemplate/sentapp.html',context)

def textgen(request):
    if request.method == 'POST':
        if 'searchinput' in request.POST:    
            searchinput = request.POST["searchinput"]
        if 'wordinput' in request.POST: 
            wordinput = request.POST["wordinput"]
        if 'lineinput' in request.POST: 
            lineinput = request.POST['lineinput']
        context={
        'searchinput':searchinput,
        'wordinput':wordinput,
        'lineinput':lineinput,
       
        
        }
        return render(request,'findtexttemplate/textgen.html',context)
    return render(request,'findtexttemplate/textgen.html')
            
def getqa(raw_text,question):
    # model = TFAutoModelForQuestionAnswering.from_pretrained(model_checkpoint)
    # question_answerer = pipeline('question-answering',model='distilbert-base-cased-distilled-squad')
    answer = question_answerer({
                'question': question,
                'context': raw_text
            })
    result=answer['answer']
    score=answer['score']
    return result,score
    


def dbsearch(request):
    dbmodel=inputtext.objects.all()
    if request.method == 'POST':
        tablename = request.POST["tabledb"]
        searchinput = request.POST["searchinput"]
        models=inputtext.objects.filter(fieldname = tablename).first()
        ltext=models.text
        # print(ltext)
        matchwords=matchword(ltext,searchinput)
        context={'dbmodel':dbmodel,
                 'matchwords':matchwords}
        return render(request,'findtexttemplate/dbsearch.html',context)
    
    return render(request,'findtexttemplate/dbsearch.html',{'dbmodel':dbmodel})



def matchword(corpus,searchinput):
    # corpus=remov_punc(corpus)
    matchwords=[]
    store=word_tokenize(corpus)
    print('store',store)
    hash_code=get_soundex(searchinput)
    token_store={}
    for token in store:
        token_store[token]=get_soundex(token)  
    print(token_store)
    for i in token_store:
        # print(token_store[i]==hash_code)
        if(token_store[i]==hash_code): 
            print(i)
            matchwords.append(i)
    print("matchwords",matchwords)
    return matchwords


def get_soundex(token):
    """Get the soundex code for the string"""
    token = token.upper()
    soundex = ""
    # first letter of input is always the first letter of soundex
    soundex += token[0]
    # create a dictionary which maps letters to respective soundex codes. Vowels and 'H', 'W' and 'Y' will be represented by '.'
    dictionary = {"BFPV": "1", "CGJKQSXZ":"2", "DT":"3", "L":"4", "MN":"5", "R":"6", "AEIOUHWY":"."}
    for char in token[1:]:
        for key in dictionary.keys():
            if char in key:
                code = dictionary[key]
                if code != soundex[-1]:
                    soundex += code
    # remove vowels and 'H', 'W' and 'Y' from soundex
    soundex = soundex.replace(".", "")
    # trim or pad to make soundex a 4-character code
    soundex = soundex[:4].ljust(4, "0")  
    return soundex


def call_fileFormat(file):
    sfile=str(file)
    ext=file.split(".")[1]
    exten="."+ext
    print(exten)
    
    if exten==".txt":
        text=read_txt(file) #call txt_file function. 
    elif exten==".xlsx":
        text=read_excel(sfile)  #call excel_file function.
    elif exten==".csv":
        text=read_csv(sfile) #call csv_file function.
    elif exten == '.pdf':
        text=extract_text_from_pdf(sfile)
    elif exten == '.docx':
        text = extract_text_from_docx(sfile)
    texts=str(text)
    
    # text=remov_punc(texts)
    print(text)
    return text


def remov_punc(text):
    punc = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    for ele in text:
        if ele in punc:
            text = text.replace(ele, "")
    return text


def read_excel(file):
    import openpyxl
    text=[]
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(file)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # Iterate the loop to read the cell values
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
#             print(col[row].value)
            val=col[row].value
            text.append(val)
    return text



def read_txt(file):

    my_file = open(file)
    data = my_file.read()
    data_into_list = data.replace('\n', ' ').split(" ")
#     print(data_into_list)
    my_file.close()
    return data_into_list


def read_csv(file):
    import csv
    with open(file,encoding='utf-8-sig') as csvfile:
        reader=csv.reader(csvfile)
        li=[]
        count=0
        i=0
        for row in reader:
            count=len(row)
            break
        for row in reader:
            i=0
            while i<count:
                li.append(row[i])
                i=i+1
        return li
    
def extract_text_from_pdf(pdf_path):
	resource_manager = PDFResourceManager()
	fake_file_handle = io.StringIO()
	converter = TextConverter(resource_manager, fake_file_handle, laparams=LAParams())
	page_interpreter = PDFPageInterpreter(resource_manager, converter)

	with open(pdf_path, 'rb') as fh:
		for page in PDFPage.get_pages(fh,
									  caching=True,
									  check_extractable=True):
			page_interpreter.process_page(page)

		text = fake_file_handle.getvalue()

	# close open handles
	converter.close()
	fake_file_handle.close()
	return text

def extract_text_from_docx(doc_path):

    try:
        temp = docx2txt.process(doc_path)
        text = [line.replace('\t', ' ') for line in temp.split('\n') if line]
        return ' '.join(text)
    except KeyError:
        return ' '
